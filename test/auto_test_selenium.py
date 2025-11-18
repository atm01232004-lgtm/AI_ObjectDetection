import unittest
import time
import os
from datetime import datetime
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from webdriver_manager.chrome import ChromeDriverManager

# Thư viện xử lý Excel
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

# --- CẤU HÌNH ---
BASE_URL = "http://127.0.0.1:5000"
TIME_WAIT = 2.0
REPORT_FOLDER = "AutoTest_Results"
REPORT_FILE = "AutoTest_Results/Bao_cao_test.xlsx"

class TestLoginFlow(unittest.TestCase):

    def setUp(self):
        print("\n>>> Đang khởi động trình duyệt...")
        chrome_options = Options()
        # chrome_options.add_argument("--headless")

        self.driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=chrome_options)
        self.driver.maximize_window()

        # Danh sách lưu kết quả test
        self.test_logs = []
        self.log_step("BẮT ĐẦU", "Khởi tạo môi trường test", "PASS")

    def tearDown(self):
        print(f"\n>>> Đang xuất báo cáo ra file Excel: {REPORT_FILE}")
        self.export_report_to_excel()

        print(">>> Đóng trình duyệt sau 3 giây.")
        time.sleep(3)
        self.driver.quit()

    def log_step(self, step_name, details, status, error_msg=""):
        # Sửa dòng này: Thêm %Y-%m-%d vào định dạng
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        self.test_logs.append([timestamp, step_name, details, status, error_msg])

        icon = "✅" if status == "PASS" else "❌"
        print(f"[{timestamp}] {icon} {step_name}: {status}")

    # --- HÀM MỚI: XUẤT RA EXCEL VỚI TÊN FILE CÓ NGÀY GIỜ ---
    def export_report_to_excel(self):
        try:
            # 1. Tạo Workbook mới
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Kết quả Auto Test"

            # 2. Tạo Tiêu đề cột
            headers = ["Thời gian", "Tên bước", "Chi tiết", "Trạng thái", "Ghi chú / Lỗi"]
            ws.append(headers)

            # 3. Trang trí Tiêu đề
            header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
            header_font = Font(bold=True)

            for col in range(1, 6):
                cell = ws.cell(row=1, column=col)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")

                # 4. Ghi dữ liệu và Tô màu
            pass_font = Font(color="008000", bold=True)
            fail_font = Font(color="FF0000", bold=True)

            for row_data in self.test_logs:
                ws.append(row_data)
                current_row = ws.max_row
                status_cell = ws.cell(row=current_row, column=4)

                if status_cell.value == "PASS":
                    status_cell.font = pass_font
                elif status_cell.value == "FAIL":
                    status_cell.font = fail_font

                # 5. Chỉnh độ rộng cột
            ws.column_dimensions['A'].width = 20  # Tăng độ rộng cột thời gian
            ws.column_dimensions['B'].width = 20
            ws.column_dimensions['C'].width = 40
            ws.column_dimensions['D'].width = 15
            ws.column_dimensions['E'].width = 30

            # 6. TẠO TÊN FILE TỰ ĐỘNG THEO NGÀY GIỜ
            # Lấy thời gian hiện tại: Năm-Tháng-Ngày_Giờ-Phút-Giây
            # Ví dụ: Ket_qua_Test_2023-10-25_14-30-05.xlsx
            current_time_str = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
            file_name = f"Ket_qua_Test_{current_time_str}.xlsx"

            # Kiểm tra xem thư mục REPORT_FOLDER đã có chưa, nếu chưa thì tạo mới
            if not os.path.exists(REPORT_FOLDER):
                os.makedirs(REPORT_FOLDER)
                print(f">>> Đã tạo thư mục mới: {REPORT_FOLDER}")

            # Ghép tên thư mục và tên file thành đường dẫn hoàn chỉnh
            # os.path.join giúp code chạy đúng trên cả Windows (\) và Mac/Linux (/)
            full_path = os.path.join(REPORT_FOLDER, file_name)

            # Lưu file vào đường dẫn mới
            wb.save(full_path)

            # Lấy đường dẫn tuyệt đối để in ra cho dễ nhìn
            abs_path = os.path.abspath(full_path)
            print(f">>> Đã lưu báo cáo thành công tại: {abs_path}")

        except Exception as e:
            print(f"Lỗi khi ghi file Excel: {e}")

    # --- KỊCH BẢN TEST (Logic cũ giữ nguyên) ---
    def test_full_flow(self):
        driver = self.driver
        try:
            # BƯỚC 1: Vào Login
            driver.get(f"{BASE_URL}/login")
            time.sleep(TIME_WAIT)
            if "Đăng nhập" in driver.title:
                self.log_step("Truy cập", "Vào trang login thành công", "PASS")
            else:
                self.log_step("Truy cập", "Sai tiêu đề trang", "FAIL")

            # BƯỚC 2: Sang Đăng Ký
            try:
                driver.find_element(By.ID, "signUp").click()
                time.sleep(TIME_WAIT)
                self.log_step("Điều hướng", "Chuyển sang màn hình Đăng Ký", "PASS")
            except Exception as e:
                self.log_step("Điều hướng", "Lỗi nút Đăng Ký", "FAIL", str(e))
                return

            # BƯỚC 3: Đăng Ký
            test_user = f"user_{int(time.time())}"
            test_pass = "123456"

            driver.find_element(By.CSS_SELECTOR, ".sign-up-container input[name='username']").send_keys(test_user)
            time.sleep(0.5)
            driver.find_element(By.CSS_SELECTOR, ".sign-up-container input[name='password']").send_keys(test_pass)
            time.sleep(0.5)
            driver.find_element(By.CSS_SELECTOR, ".sign-up-container button").click()

            self.log_step("Đăng Ký", f"Đã tạo user: {test_user}", "PASS")
            time.sleep(TIME_WAIT)

            # BƯỚC 4: Đăng Nhập (Trang tự reload)
            print(f"--- Trang tự reload về Login ---")
            try:
                u_in = driver.find_element(By.CSS_SELECTOR, ".sign-in-container input[name='username']")
                p_in = driver.find_element(By.CSS_SELECTOR, ".sign-in-container input[name='password']")
                b_in = driver.find_element(By.CSS_SELECTOR, ".sign-in-container button")

                u_in.clear()
                u_in.send_keys(test_user)
                time.sleep(0.5)
                p_in.send_keys(test_pass)
                time.sleep(0.5)
                b_in.click()
                self.log_step("Đăng Nhập", "Đã điền thông tin và submit", "PASS")
            except Exception as e:
                self.log_step("Đăng Nhập", "Không tìm thấy form", "FAIL", str(e))
                return

            time.sleep(TIME_WAIT)

            # BƯỚC 5: Kiểm tra Dashboard
            if driver.current_url == f"{BASE_URL}/":
                self.log_step("Kết quả URL", "Đúng trang chủ", "PASS")
            else:
                self.log_step("Kết quả URL", f"Sai URL: {driver.current_url}", "FAIL")

            if test_user in driver.page_source:
                self.log_step("Hiển thị User", f"Thấy user '{test_user}'", "PASS")
            else:
                self.log_step("Hiển thị User", "Không thấy tên user", "FAIL")

        except Exception as e:
            self.log_step("LỖI HỆ THỐNG", "Chương trình dừng đột ngột", "FAIL", str(e))


if __name__ == "__main__":
    unittest.main()