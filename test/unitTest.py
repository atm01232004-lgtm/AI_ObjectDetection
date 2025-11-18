import unittest
import sys
import os
from datetime import datetime
from app import app, users_db

# Thư viện Excel
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

# --- CẤU HÌNH TÊN THƯ MỤC ---
REPORT_FOLDER = "Result_UnitTest"


class TestAILoginSystem(unittest.TestCase):
    # Biến lưu trữ kết quả dùng chung cho toàn bộ Class
    test_results = []

    # --- PHẦN 1: CẤU HÌNH MÔI TRƯỜNG ---
    def setUp(self):
        """Chạy TRƯỚC mỗi bài test"""
        app.config['TESTING'] = True
        app.config['WTF_CSRF_ENABLED'] = False
        app.config['SECRET_KEY'] = 'test_secret_key'
        self.client = app.test_client()

        users_db.clear()
        users_db['admin'] = '123456'

    def tearDown(self):
        """Chạy SAU mỗi bài test: Bắt kết quả PASS/FAIL"""
        method_id = self.id().split('.')[-1]
        description = self._testMethodDoc or "Không có mô tả"

        # Logic xác định PASS/FAIL
        result_list = self._outcome.errors
        status = "PASS"
        error_msg = ""

        for test, exc_info in result_list:
            if exc_info:
                status = "FAIL"
                error_msg = str(exc_info[1])

                # Thêm Ngày-Giờ vào từng dòng log
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        self.test_results.append([timestamp, method_id, description, status, error_msg])

    # --- PHẦN 2: XUẤT EXCEL VÀO FOLDER (QUAN TRỌNG) ---
    @classmethod
    def tearDownClass(cls):
        """Chạy sau khi TẤT CẢ test xong -> Lưu file vào folder Result_UnitTest"""
        print(f"\n>>> Đang xuất báo cáo vào thư mục: {REPORT_FOLDER}...")

        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Unit Test Report"

            # 1. Tạo tiêu đề cột
            headers = ["Thời gian", "Tên Test Case", "Mô tả chức năng", "Trạng thái", "Chi tiết lỗi"]
            ws.append(headers)

            # 2. Trang trí tiêu đề
            header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            header_font = Font(bold=True)

            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")

            # 3. Ghi dữ liệu
            pass_font = Font(color="008000", bold=True)
            fail_font = Font(color="FF0000", bold=True)

            for row_data in cls.test_results:
                ws.append(row_data)
                current_row = ws.max_row

                status_cell = ws.cell(row=current_row, column=4)
                if status_cell.value == "PASS":
                    status_cell.font = pass_font
                else:
                    status_cell.font = fail_font

            # 4. Chỉnh độ rộng cột
            ws.column_dimensions['A'].width = 20
            ws.column_dimensions['B'].width = 30
            ws.column_dimensions['C'].width = 40
            ws.column_dimensions['D'].width = 12
            ws.column_dimensions['E'].width = 40

            # --- ĐOẠN CODE TẠO FOLDER VÀ LƯU FILE ---

            # Kiểm tra và tạo thư mục nếu chưa có
            if not os.path.exists(REPORT_FOLDER):
                os.makedirs(REPORT_FOLDER)
                print(f">>> Đã tạo thư mục mới: {REPORT_FOLDER}")

            # Tạo tên file theo ngày giờ: Report_2023-11-17_10-30-00.xlsx
            current_time_str = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
            file_name = f"Report_{current_time_str}.xlsx"

            # Ghép đường dẫn: Result_UnitTest/Report_...xlsx
            full_path = os.path.join(REPORT_FOLDER, file_name)

            wb.save(full_path)
            print(f">>> Đã lưu thành công tại: {os.path.abspath(full_path)}")

        except Exception as e:
            print(f"Lỗi khi xuất Excel: {e}")

    # --- PHẦN 3: CÁC TEST CASE (Giữ nguyên) ---

    def test_login_page_loads(self):
        """Kiểm tra trang đăng nhập có tải lên bình thường không"""
        response = self.client.get('/login')
        self.assertEqual(response.status_code, 200)
        self.assertIn(b'\xc4\x90\xc4\x83ng Nh\xe1\xba\xadp', response.data)
        self.assertIn(b'name="username"', response.data)

    def test_home_redirects_when_not_logged_in(self):
        """Chưa đăng nhập mà vào trang chủ thì phải bị đá về login"""
        response = self.client.get('/', follow_redirects=True)
        self.assertIn(b'\xc4\x90\xc4\x83ng Nh\xe1\xba\xadp', response.data)

    def test_register_success(self):
        """Test đăng ký tài khoản mới thành công"""
        response = self.client.post('/login', data=dict(
            username='newuser',
            password='password123',
            action='register'
        ), follow_redirects=True)
        self.assertIn(b'\xc4\x90\xc4\x83ng k\xc3\xbd th\xc3\xa0nh c\xc3\xb4ng', response.data)
        self.assertIn('newuser', users_db)

    def test_register_duplicate(self):
        """Test đăng ký trùng tên admin đã có"""
        response = self.client.post('/login', data=dict(
            username='admin',
            password='newpass',
            action='register'
        ), follow_redirects=True)
        self.assertIn(b'T\xc3\xa0i kho\xe1\xba\xa3n \xc4\x91\xc3\xa3 t\xe1\xbb\x93n t\xe1\xba\xa1i', response.data)

    def test_login_success(self):
        """Test đăng nhập đúng"""
        response = self.client.post('/login', data=dict(
            username='admin',
            password='123456',
            action='login'
        ), follow_redirects=True)
        self.assertIn(b'AI Vision Pro', response.data)

    def test_login_failure(self):
        """Test đăng nhập sai mật khẩu"""
        response = self.client.post('/login', data=dict(
            username='admin',
            password='wrongpassword',
            action='login'
        ), follow_redirects=True)
        self.assertIn(b'Sai t\xc3\xa0i kho\xe1\xba\xa3n ho\xe1\xba\xb7c m\xe1\xba\xadt kh\xe1\xba\xa9u', response.data)

    def test_logout(self):
        """Test chức năng đăng xuất"""
        self.client.post('/login', data=dict(username='admin', password='123456', action='login'),
                         follow_redirects=True)
        response = self.client.get('/logout', follow_redirects=True)
        self.assertIn(b'\xc4\x90\xc4\x83ng Nh\xe1\xba\xadp', response.data)


if __name__ == '__main__':
    unittest.main()